"""Property-attribute reference: fills Property Type, Bedroom and Size by location.

WHY THIS EXISTS
Property Type is present in roughly 40% of source rows, and where it is present
it is stated in registry vocabulary ("Unit", "Flat", "Land") rather than market
vocabulary. cleaning.clean_property_type() fixes the vocabulary; it cannot
invent the missing 60%.

Structural inference was measured against the real registers before this module
was written, and it does not work. FLOOR -- the one signal that reliably
separates an apartment from a villa -- is populated in 0% of sampled rows, and
the remaining combinations (unit number, plot number, building, bedroom) are
spread almost evenly between property kinds in a mixed-use master community.
Guessing from them would repeat the mistake this codebase already made once,
when development categories were written into Property Type and made records
look complete while being wrong.

So the missing 60% has to come from data that actually knows the answer: a
property dataset keyed by location, such as a portal export. This module loads
one and matches at three precisions, most specific first:

    community | building | unit   the exact property        -- authoritative
    community | building          dominant type in a tower  -- a tower is
                                                               almost always
                                                               one kind
    community                     dominant type in an area  -- only when the
                                                               area is lopsided

Each precision has its own dominance threshold, and a group that fails its
threshold fills nothing. A community that is 55% villas must not stamp "Villa"
on the other 45%.

INPUT FORMAT
Any CSV or Excel export with a header row. Columns are matched by name, and the
recognised spellings are listed in _COLUMN_ALIASES below, because Property
Finder, Property Monitor, Bayut and DLD exports all use different ones. Only a
location and a property type are required:

    community, building (or tower/project), property_type
    optional: unit_number, bedrooms, size

Nothing here scrapes anything. It reads a file you supply, whatever its origin.
"""
from __future__ import annotations

import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from . import cleaning as C

log = logging.getLogger("engine.property_reference")

# Column spellings seen across portal and registry exports. Matched after
# normalising to uppercase with punctuation collapsed.
_COLUMN_ALIASES = {
    "community": ("COMMUNITY", "MASTER COMMUNITY", "MASTER PROJECT",
                  "MASTER PROJECT EN", "AREA NAME EN", "COMMUNITY EN",
                  "NEIGHBOURHOOD", "LOCATION", "AREA"),
    "building": ("BUILDING", "TOWER", "BUILDING NAME", "TOWER NAME",
                 "BUILDING NAME EN", "CLUSTER", "SUB COMMUNITY",
                 "SUBCOMMUNITY", "PROJECT NAME", "PROJECT EN", "PROJECT",
                 "DEVELOPMENT"),
    "unit_number": ("UNIT NUMBER", "UNIT NO", "UNITNUMBER", "UNIT NUMBER EN",
                    "PROPERTY NUMBER", "UNIT"),
    "property_type": ("PROPERTY TYPE", "PROPERTY TYPE EN", "PROPERTYTYPE",
                      "PROPERTY SUB TYPE", "PROPERTY SUBTYPE", "UNIT TYPE",
                      "CATEGORY", "TYPE"),
    "bedrooms": ("BEDROOMS", "BEDROOM", "BEDS", "NO OF BEDROOMS", "ROOMS EN",
                 "BR"),
    "size": ("SIZE", "SIZE SQFT", "AREA SQFT", "BUILT UP AREA", "BUA",
             "PROCEDURE AREA", "ACTUAL AREA"),
    # Portal exports identify a property by one composite location string
    # rather than separate columns; see _parse_location().
    "location_full_name": ("LOCATION FULL NAME", "FULL LOCATION", "ADDRESS"),
    "listing_id": ("ID", "LISTING ID", "PROPERTY ID", "REFERENCE"),
}

# Trailing segment of a portal location string: the city, never part of the
# community name. Verified against 90,807 Property Finder rows, all of which
# end in "Dubai".
_CITIES = {"DUBAI", "ABU DHABI", "SHARJAH", "AJMAN", "FUJAIRAH",
           "RAS AL KHAIMAH", "UMM AL QUWAIN", "UAE"}


def _parse_location(full_name) -> tuple[str, str, str]:
    """Split a portal location string into (community, building, sub_community).

    Property Finder writes location as a comma-separated path from the most
    specific part to the least:

        "Skycourts Tower F, Skycourts Towers, Dubai Land Residence Complex, Dubai"
         ^ building        ^ sub-community    ^ community                    ^ city

    The city is always last, so the community is the segment before it and the
    building is the first. Anything between is a sub-community, which is worth
    keeping because owner registers put that name in Building/Cluster at least
    as often as they put the tower name there.
    """
    segs = [s.strip() for s in str(full_name or "").split(",") if s.strip()]
    if segs and segs[-1].upper() in _CITIES:
        segs = segs[:-1]
    if not segs:
        return "", "", ""
    if len(segs) == 1:
        return segs[0], "", ""              # community only, no building named
    community = segs[-1]
    building = segs[0]
    sub_community = segs[1] if len(segs) >= 3 else ""
    return community, building, sub_community

# A dominant type is trusted only when it holds this share of a group and the
# group has at least this many properties. The bar rises as the location gets
# broader, because a whole community is far likelier to be mixed-use than a
# single tower.
DOMINANCE = {
    "building": (0.80, 3),
    "community": (0.90, 25),
}


def _norm_header(h) -> str:
    return C._WS_RE.sub(" ", C._PT_PUNCT_RE.sub(" ", str(h or "").upper())).strip()


def _key(*parts) -> str:
    return "|".join((p or "").strip().upper() for p in parts)


# Portal listings name a sub-development more fully than an owner register does:
# "Maple at Dubai Hills Estate" vs "Maple", "Sidra Villas" vs "Sidra",
# "Shoreline Apartments" vs "Shoreline". Measured against the real files, this
# naming gap -- not missing data -- was what stopped building-level matches from
# firing on the largest communities. Both sides are reduced to the same key.
_AT_SUFFIX_RE = re.compile(r"\s+\bat\b\s+.*$", re.I)
_PHASE_SUFFIX_RE = re.compile(r"\s+\b(?:phase|ph|cluster|building|bldg)\b\.?\s*\d*$", re.I)
# Generic descriptors that a register omits and a portal spells out. Stripped
# only from the END, so "Villa Lantana" keeps its leading word.
_GENERIC_TAIL_RE = re.compile(
    r"\s+\b(?:villas?|residences?|apartments?|towers?|homes?|estate|"
    r"community|complex|development|project)\b$", re.I)
_BUILDING_PUNCT_RE = re.compile(r"[^a-z0-9 ]+")


def _building_key(name) -> str:
    """Reduce a building / sub-community name to a form both sides share."""
    s = str(name or "").strip().lower()
    if not s:
        return ""
    s = _AT_SUFFIX_RE.sub("", s)
    s = _PHASE_SUFFIX_RE.sub("", s)
    # Repeat: "Sidra Villas Residences" sheds one descriptor per pass.
    for _ in range(3):
        stripped = _GENERIC_TAIL_RE.sub("", s)
        if stripped == s:
            break
        s = stripped
    s = _BUILDING_PUNCT_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Reducing a name to nothing means it was purely generic ("The Towers");
    # keep the original so it still keys something distinct.
    return s or _BUILDING_PUNCT_RE.sub(" ", str(name or "").strip().lower()).strip()


@dataclass(frozen=True)
class PropertyFacts:
    property_type: str | None = None
    bedroom: str | None = None
    size: float | None = None
    precision: str = "unit"          # unit | building | community
    support: int = 1                 # how many source properties backed this


class PropertyReference:
    """Location -> property attributes, at three precisions."""

    def __init__(self, rows: list[dict]):
        self.rows = rows
        self._by_unit: dict[str, PropertyFacts] = {}
        by_building: dict[str, Counter] = defaultdict(Counter)
        by_community: dict[str, Counter] = defaultdict(Counter)

        for r in rows:
            comm = C.clean_community(r.get("community")) or ""
            bldg = C.clean_text(r.get("building")) or ""
            unit = C.clean_unit(r.get("unit_number")) or ""
            ptype = C.clean_property_type(r.get("property_type"))
            if not ptype or not comm:
                continue

            # Bedroom and size are carried at unit precision only. A "typical"
            # bedroom count for a tower is not a fact about any apartment in it.
            if unit:
                self._by_unit[_key(comm, bldg, unit)] = PropertyFacts(
                    property_type=ptype,
                    bedroom=C.clean_bedroom(r.get("bedrooms"))[0],
                    size=C.clean_size(r.get("size")),
                    precision="unit",
                )
            # Indexed under the tower name AND the sub-community, because an
            # owner register puts either one in Building/Cluster depending on
            # which the builder happened to use. Both describe the same set of
            # properties, so both are legitimate keys for the same counts.
            sub = C.clean_text(r.get("sub_community")) or ""
            for name in {_building_key(n) for n in (bldg, sub) if n}:
                if name:
                    by_building[_key(comm, name)][ptype] += 1
            by_community[_key(comm)][ptype] += 1

        self._by_building = self._distil(by_building, "building")
        self._by_community = self._distil(by_community, "community")
        # Community-agnostic index over the same building / sub-community
        # names. Owner registers frequently carry no Community at all (76% of
        # sampled rows) and, when they do, they name a sub-development
        # ("Club Villas") where the portal names the master community ("Dubai
        # Hills Estate"). Keying the development name on its own is what makes
        # those rows matchable; the dominance threshold still applies, so a
        # name that means different things in different places matches nothing.
        by_name: dict[str, Counter] = defaultdict(Counter)
        for key, counts in by_building.items():
            name = key.split("|", 1)[1] if "|" in key else key
            if name:
                by_name[name].update(counts)
        self._by_name = self._distil(by_name, "building")

    @staticmethod
    def _distil(counters, precision) -> dict[str, PropertyFacts]:
        """Keep only groups lopsided enough for one type to speak for them."""
        share_min, count_min = DOMINANCE[precision]
        out: dict[str, PropertyFacts] = {}
        for key, counts in counters.items():
            total = sum(counts.values())
            if total < count_min:
                continue
            ptype, n = counts.most_common(1)[0]
            if n / total >= share_min:
                out[key] = PropertyFacts(property_type=ptype,
                                         precision=precision, support=total)
        return out

    def lookup(self, community, building, unit) -> PropertyFacts | None:
        comm = C.clean_community(community) or ""
        bldg = C.clean_text(building) or ""
        if not comm and not bldg and not community:
            return None
        un = C.clean_unit(unit) or ""

        if un:
            hit = self._by_unit.get(_key(comm, bldg, un))
            if hit:
                return hit
        if bldg:
            hit = self._by_building.get(_key(comm, _building_key(bldg)))
            if hit:
                return hit

        # Development name alone, community ignored. Tried for the building
        # field and for the community field, because a register's "Community"
        # is as often a sub-development as a master community.
        for candidate in (bldg, community):
            name = _building_key(candidate)
            if name:
                hit = self._by_name.get(_key(name))
                if hit:
                    return hit

        return self._by_community.get(_key(comm))

    def stats(self) -> dict:
        return {
            "source_rows": len(self.rows),
            "units": len(self._by_unit),
            "buildings": len(self._by_building),
            "communities": len(self._by_community),
        }

    def __len__(self) -> int:
        return len(self._by_unit) + len(self._by_building) + len(self._by_community)


def _resolve_columns(header: list[str]) -> dict[str, int]:
    """Map our field names onto column indexes in the supplied export."""
    norm = [_norm_header(h) for h in header]
    found: dict[str, int] = {}
    for field, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:            # alias order encodes preference
            if alias in norm:
                found[field] = norm.index(alias)
                break
    return found


def _iter_sheets(path: Path):
    """Yield (header, row_iterator) for a CSV or Excel export.

    Deliberately NOT engine.detection.open_source(). That module hunts for a
    header row by matching against the 23 owner-register field names, and a
    portal export (id, property_type, title, price_aed, ...) looks like nothing
    it recognises -- it reports the file as headerless and every column is lost.
    A property export is a different kind of file with a plain header on row 1,
    so it is read as one.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv
        # newline="" per the csv module contract; utf-8-sig drops the BOM Excel
        # writes, which would otherwise corrupt the first column name.
        with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
            reader = csv.reader(fh)
            try:
                header = next(reader)
            except StopIteration:
                return
            yield header, reader
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl not available; skipping %s", path)
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            try:
                header = ["" if c is None else str(c) for c in next(it)]
            except StopIteration:
                continue
            yield header, it
    finally:
        wb.close()


def _read_file(path: Path, rows: list[dict], seen_ids: set) -> None:
    """Append one export's rows to `rows`, skipping ids already collected."""
    for header, row_iter in _iter_sheets(path):
        cols = _resolve_columns(header)
        # Either explicit community/building columns, or one composite portal
        # location string that carries both.
        has_location = "community" in cols or "location_full_name" in cols
        if "property_type" not in cols or not has_location:
            log.debug("skipping %s: no usable columns", path.name)
            continue

        for raw in row_iter:
            def cell(field):
                i = cols.get(field)
                return raw[i] if i is not None and i < len(raw) else None

            # Exports are often delivered as overlapping partial dumps; the id
            # is what makes re-reading all of them idempotent.
            listing_id = cell("listing_id")
            if listing_id not in (None, ""):
                marker = (path.name if "location_full_name" not in cols
                          else "portal", str(listing_id))
                if marker[1] in seen_ids:
                    continue
                seen_ids.add(marker[1])

            if "location_full_name" in cols:
                community, building, sub_community = _parse_location(
                    cell("location_full_name"))
            else:
                community = cell("community")
                building = cell("building")
                sub_community = None

            rows.append({
                "community": community,
                "building": building,
                "sub_community": sub_community,
                "unit_number": cell("unit_number"),
                "property_type": cell("property_type"),
                "bedrooms": cell("bedrooms"),
                "size": cell("size"),
            })


def load_property_reference(path: str | Path) -> PropertyReference:
    """Read a property export into a PropertyReference.

    `path` may be a single CSV/Excel file or a directory of them. A directory
    is read whole and de-duplicated on listing id, because portal scrapes tend
    to arrive as several overlapping dumps rather than one clean file.

    Returns an empty reference rather than raising: a missing or malformed
    property file must degrade enrichment, never fail an ingest.
    """
    path = Path(path)
    if not path.exists():
        log.info("property reference not found: %s", path)
        return PropertyReference([])

    rows: list[dict] = []
    seen_ids: set = set()
    if path.is_dir():
        files = sorted(
            [p for p in path.iterdir()
             if p.suffix.lower() in (".csv", ".xlsx", ".xls")],
            # Largest first: the most complete dump seeds the id set, so the
            # partial ones add only what it is missing.
            key=lambda p: -p.stat().st_size,
        )
    else:
        files = [path]

    for f in files:
        _read_file(f, rows, seen_ids)

    ref = PropertyReference(rows)
    log.info("property reference loaded from %s: %s", path, ref.stats())
    return ref
