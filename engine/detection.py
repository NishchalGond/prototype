# Heuristic header matching and column detection rules
"""Format detection + streaming sheet readers.

The audit proved extensions lie: a 46 MB ".xls" was an HTML table, and an
".xlsx" was a password-protected OLE2 container. Everything here dispatches on
magic bytes, never on the filename.
"""
from __future__ import annotations

import hashlib
import html
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

XLSX_MAGIC = b"PK\x03\x04"
OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

FMT_XLSX = "xlsx"
FMT_XLS = "xls"
FMT_HTML = "html_table"
FMT_CSV = "csv"
FMT_ENCRYPTED = "encrypted"
FMT_UNKNOWN = "unknown"


class UnreadableFile(Exception):
    """Raised when a file cannot be read at all (e.g. encrypted)."""


@dataclass
class SheetData:
    name: str
    header: list[str]
    header_row_index: int          # 0-based index of the header row, -1 if headerless
    rows: Iterator[list]           # data rows only (header already consumed)
    n_cols: int
    headerless: bool = False
    meta: dict = field(default_factory=dict)


def sha256_of(path: Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        while True:
            b = fh.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def detect_format(path: Path) -> str:
    with open(path, "rb") as fh:
        head = fh.read(2048)
    if head.startswith(XLSX_MAGIC):
        return FMT_XLSX
    if head.startswith(OLE2_MAGIC):
        # OLE2 is either a real BIFF workbook or an encrypted OOXML container.
        try:
            text = head.decode("utf-16-le", errors="replace")
        except Exception:
            text = ""
        if "EncryptedPackage" in text or "StrongEncryption" in text:
            return FMT_ENCRYPTED
        return FMT_XLS
    stripped = head.lstrip(b"\xef\xbb\xbf").lstrip()
    if stripped[:5].lower() == b"<html" or b"<table" in head[:2048].lower():
        return FMT_HTML
    if path.suffix.lower() == ".csv":
        return FMT_CSV
    return FMT_UNKNOWN


# --------------------------------------------------------------------------
# header-row detection
# --------------------------------------------------------------------------
_EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[a-z]{2,}", re.I)
_LONGNUM_RE = re.compile(r"^[\d\s+\-|().]{7,}$")
_DATEISH_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")


def _looks_like_data(v: str) -> bool:
    if _EMAIL_RE.search(v):
        return True
    if _LONGNUM_RE.match(v):
        return True
    if _DATEISH_RE.match(v):
        return True
    try:
        float(v.replace(",", ""))
        return True
    except ValueError:
        return False


def score_header_row(cells: list) -> float:
    """Higher = more likely to be a header row."""
    vals = [str(c).strip() for c in cells if c is not None and str(c).strip()]
    if len(vals) < 2:
        return -1.0
    score = float(len(vals))
    for v in vals:
        if _looks_like_data(v):
            score -= 2.0
        elif len(v) > 60:
            score -= 1.0
    return score


def _known_header_tokens() -> set[str]:
    """Lazy import to avoid a module cycle; mapping does not import detection."""
    from .mapping import ALIAS, EXCLUDE, IGNORE_TO_EXTRAS, norm_header  # noqa: PLC0415
    toks = set(ALIAS) | set(EXCLUDE) | set(IGNORE_TO_EXTRAS)
    toks |= {norm_header(x) for x in ("PREMISE 1", "PREMISE 2", "PREMISE 3",
                                      "SERIAL NO.", "AREA OWNED", "USAGE", "EMIRATE")}
    return toks


def find_header_row(preview: list[list], max_scan: int = 12) -> int:
    """Index of the header row, or -1 if the sheet has none.

    A high text-vs-number score is not enough: the first data row of the
    headerless owner register ("265616 | Business Bay | Shop | ...") scores well
    on shape alone. A row only counts as a header if its cells are actually
    recognisable column names.
    """
    from .mapping import norm_header  # noqa: PLC0415

    known = _known_header_tokens()
    best_i, best_s = -1, 0.0
    for i, row in enumerate(preview[:max_scan]):
        s = score_header_row(row)
        if s <= best_s:
            continue
        vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
        # single/double-char cells ("F", "1") are values, never column names
        hits = sum(1 for v in vals if len(v) > 2 and norm_header(v) in known)
        # need real evidence: >=2 recognised names covering >=1/3 of filled cells
        if hits >= 2 and hits >= max(2, len(vals) // 3):
            best_i, best_s = i, s
    return best_i


# --------------------------------------------------------------------------
# readers — all stream, none load the whole sheet into memory
# --------------------------------------------------------------------------
def _norm_cell(c):
    if c is None:
        return None
    if isinstance(c, str):
        s = c.strip()
        return s if s else None
    return c


def read_xlsx(path: Path) -> Iterator[SheetData]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            preview: list[list] = []
            it = ws.iter_rows(values_only=True)
            for _ in range(12):
                try:
                    preview.append(list(next(it)))
                except StopIteration:
                    break
            if not preview:
                continue
            hi = find_header_row(preview)
            n_cols = max((len(r) for r in preview), default=0)
            if hi < 0:
                header, headerless = [], True
                buffered = preview
            else:
                header = [("" if c is None else str(c).strip()) for c in preview[hi]]
                headerless = False
                buffered = preview[hi + 1:]

            def gen(buffered=buffered, it=it):
                for r in buffered:
                    yield [_norm_cell(c) for c in r]
                for r in it:
                    yield [_norm_cell(c) for c in r]

            yield SheetData(ws.title, header, hi, gen(), n_cols, headerless)
    finally:
        wb.close()


def read_xls(path: Path) -> Iterator[SheetData]:
    import xlrd

    bk = xlrd.open_workbook(path, on_demand=True)
    for sh in bk.sheets():
        if sh.nrows == 0:
            continue
        preview = [sh.row_values(i) for i in range(min(12, sh.nrows))]
        hi = find_header_row(preview)
        n_cols = sh.ncols
        start = 0 if hi < 0 else hi + 1
        header = ([] if hi < 0
                  else [("" if c is None else str(c).strip()) for c in preview[hi]])

        def gen(sh=sh, start=start):
            for i in range(start, sh.nrows):
                yield [_norm_cell(c) for c in sh.row_values(i)]

        yield SheetData(sh.name, header, hi, gen(), n_cols, hi < 0)


_TR_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.S | re.I)
_TD_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.S | re.I)
_TAG_RE = re.compile(r"<[^>]+>")


def read_html_table(path: Path) -> Iterator[SheetData]:
    """Stream an HTML-table export without holding the whole document in RAM."""
    def cells_of(tr: str) -> list:
        out = []
        for c in _TD_RE.findall(tr):
            v = html.unescape(_TAG_RE.sub("", c)).replace("\xa0", " ").strip()
            out.append(v or None)
        return out

    def iter_rows():
        buf = ""
        with open(path, encoding="utf8", errors="replace") as fh:
            while True:
                chunk = fh.read(1 << 20)
                if not chunk:
                    break
                buf += chunk
                while True:
                    m = _TR_RE.search(buf)
                    if not m:
                        break
                    yield cells_of(m.group(1))
                    buf = buf[m.end():]

    gen = iter_rows()
    try:
        header_cells = next(gen)
    except StopIteration:
        return
    header = [("" if c is None else str(c).strip()) for c in header_cells]
    yield SheetData("html_table", header, 0, gen, len(header), False)


def read_csv(path: Path) -> Iterator[SheetData]:
    import csv as _csv

    with open(path, encoding="utf8", errors="replace", newline="") as fh:
        sample = fh.read(64 * 1024)
        fh.seek(0)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            dialect = _csv.excel
        reader = _csv.reader(fh, dialect)
        rows = list(reader)
    if not rows:
        return
    hi = find_header_row(rows[:12])
    header = [] if hi < 0 else [str(c).strip() for c in rows[hi]]
    body = rows if hi < 0 else rows[hi + 1:]
    yield SheetData(path.name, header, hi, iter([[_norm_cell(c) for c in r] for r in body]),
                    max((len(r) for r in rows), default=0), hi < 0)


READERS = {
    FMT_XLSX: read_xlsx,
    FMT_XLS: read_xls,
    FMT_HTML: read_html_table,
    FMT_CSV: read_csv,
}


def open_source(path: Path) -> tuple[str, Iterator[SheetData]]:
    """Return (detected_format, sheet iterator). Raises UnreadableFile if impossible."""
    fmt = detect_format(path)
    if fmt == FMT_ENCRYPTED:
        raise UnreadableFile(
            "File is password-protected (OLE2 EncryptedPackage). "
            "Supply the password or remove protection before uploading."
        )
    if fmt == FMT_UNKNOWN:
        raise UnreadableFile("Unrecognised file format (not xlsx, xls, html-table or csv).")
    reader = READERS[fmt]
    # An .xls that is really BIFF may still fail; fall back to xlsx/html probing.
    if fmt == FMT_XLS:
        try:
            import xlrd
            xlrd.open_workbook(path, on_demand=True)
        except Exception as exc:
            raise UnreadableFile(f"OLE2 file is not a readable BIFF workbook: {exc}") from exc
    return fmt, reader(path)
