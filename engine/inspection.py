"""Cheap structural inspection performed at upload time.

The dashboard needs row counts and mapped-field counts *before* processing
starts, so the upload response can populate its tiles. This must not read every
row of a 46 MB file, so it uses the workbook's declared dimensions where they
exist and a fast streaming count where they do not.
"""
from __future__ import annotations

import re
from dataclasses import asdict, dataclass, field
from pathlib import Path

from .detection import (
    FMT_CSV, FMT_HTML, FMT_XLS, FMT_XLSX, UnreadableFile, detect_format,
    find_header_row,
)
from .mapping import build_plan, is_reference_sheet


@dataclass
class SheetInfo:
    name: str
    total_rows: int
    n_cols: int
    headerless: bool
    is_reference: bool
    mapped_targets: list[str] = field(default_factory=list)
    unmapped_headers: list[str] = field(default_factory=list)
    header: list[str] = field(default_factory=list)


@dataclass
class FileInfo:
    detected_format: str
    sheet_count: int
    total_rows: int
    mapped_target_count: int
    sheets: list[SheetInfo] = field(default_factory=list)

    def to_dict(self) -> dict:
        d = asdict(self)
        d["sheets"] = [asdict(s) if not isinstance(s, dict) else s for s in self.sheets]
        return d


def _plan_for(header, headerless, n_cols, samples):
    plan = build_plan(header, headerless, n_cols, samples)
    targets = sorted(set(plan.index_to_target.values()))
    # composites contribute targets too
    if "premise1" in plan.composite:
        for t in ("Community", "Building/Cluster", "Unit Number"):
            if t not in targets:
                targets.append(t)
    return targets, plan.unmapped_headers


def _inspect_xlsx(path: Path) -> FileInfo:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[SheetInfo] = []
    try:
        for ws in wb.worksheets:
            preview = []
            for _, row in zip(range(30), ws.iter_rows(max_row=30, values_only=True)):
                preview.append(list(row))
            if not preview:
                continue
            hi = find_header_row(preview)
            n_cols = max((len(r) for r in preview), default=0)
            header = ([] if hi < 0
                      else [("" if c is None else str(c).strip()) for c in preview[hi]])
            body = preview[hi + 1:] if hi >= 0 else preview
            samples = {i: [r[i] for r in body if i < len(r)] for i in range(n_cols)}
            targets, unmapped = _plan_for(header, hi < 0, n_cols, samples)
            # read_only workbooks expose the declared dimension without a scan
            rows = max(0, (ws.max_row or 0) - (hi + 1 if hi >= 0 else 0))
            sheets.append(SheetInfo(
                name=ws.title, total_rows=rows, n_cols=n_cols, headerless=hi < 0,
                is_reference=is_reference_sheet(header, ws.title),
                mapped_targets=targets, unmapped_headers=unmapped[:20],
                header=header[:60],
            ))
    finally:
        wb.close()
    return _assemble(FMT_XLSX, sheets)


def _inspect_xls(path: Path) -> FileInfo:
    import xlrd

    bk = xlrd.open_workbook(path, on_demand=True)
    sheets: list[SheetInfo] = []
    for sh in bk.sheets():
        if sh.nrows == 0:
            continue
        preview = [sh.row_values(i) for i in range(min(30, sh.nrows))]
        hi = find_header_row(preview)
        header = ([] if hi < 0
                  else [("" if c is None else str(c).strip()) for c in preview[hi]])
        n_cols = sh.ncols
        body = preview[hi + 1:] if hi >= 0 else preview
        samples = {i: [r[i] for r in body if i < len(r)] for i in range(n_cols)}
        targets, unmapped = _plan_for(header, hi < 0, n_cols, samples)
        sheets.append(SheetInfo(
            name=sh.name, total_rows=max(0, sh.nrows - (hi + 1 if hi >= 0 else 0)),
            n_cols=n_cols, headerless=hi < 0,
            is_reference=is_reference_sheet(header, sh.name),
            mapped_targets=targets, unmapped_headers=unmapped[:20], header=header[:60],
        ))
    return _assemble(FMT_XLS, sheets)


_TR_OPEN = re.compile(rb"<tr[\s>]", re.I)


def _inspect_html(path: Path) -> FileInfo:
    from .detection import read_html_table

    # count <tr> by streaming bytes — no parsing, no full document in memory
    rows = 0
    with open(path, "rb") as fh:
        tail = b""
        while chunk := fh.read(1 << 22):
            buf = tail + chunk
            rows += len(_TR_OPEN.findall(buf))
            tail = buf[-8:]
    header: list[str] = []
    targets: list[str] = []
    unmapped: list[str] = []
    n_cols = 0
    for sheet in read_html_table(path):
        header = sheet.header
        n_cols = sheet.n_cols
        samples: dict[int, list] = {}
        for _, r in zip(range(30), sheet.rows):
            for i, v in enumerate(r):
                samples.setdefault(i, []).append(v)
        targets, unmapped = _plan_for(header, False, n_cols, samples)
        break
    info = SheetInfo(name="html_table", total_rows=max(0, rows - 1), n_cols=n_cols,
                     headerless=False, is_reference=False, mapped_targets=targets,
                     unmapped_headers=unmapped[:20], header=header[:60])
    return _assemble(FMT_HTML, [info])


def _inspect_csv(path: Path) -> FileInfo:
    from .detection import read_csv

    rows = 0
    with open(path, "rb") as fh:
        while chunk := fh.read(1 << 22):
            rows += chunk.count(b"\n")
    header: list[str] = []
    targets: list[str] = []
    unmapped: list[str] = []
    n_cols = 0
    for sheet in read_csv(path):
        header, n_cols = sheet.header, sheet.n_cols
        samples: dict[int, list] = {}
        for _, r in zip(range(30), sheet.rows):
            for i, v in enumerate(r):
                samples.setdefault(i, []).append(v)
        targets, unmapped = _plan_for(header, sheet.headerless, n_cols, samples)
        break
    return _assemble(FMT_CSV, [SheetInfo(
        name=path.name, total_rows=max(0, rows - 1), n_cols=n_cols, headerless=False,
        is_reference=False, mapped_targets=targets, unmapped_headers=unmapped[:20],
        header=header[:60])])


def _assemble(fmt: str, sheets: list[SheetInfo]) -> FileInfo:
    all_targets: set[str] = set()
    total = 0
    for s in sheets:
        if not s.is_reference:
            total += s.total_rows
            all_targets.update(s.mapped_targets)
    return FileInfo(detected_format=fmt, sheet_count=len(sheets), total_rows=total,
                    mapped_target_count=len(all_targets), sheets=sheets)


_INSPECTORS = {
    FMT_XLSX: _inspect_xlsx,
    FMT_XLS: _inspect_xls,
    FMT_HTML: _inspect_html,
    FMT_CSV: _inspect_csv,
}


def inspect_source(path: Path) -> FileInfo:
    """Structure summary for the upload response. Raises UnreadableFile."""
    fmt = detect_format(path)
    if fmt not in _INSPECTORS:
        raise UnreadableFile(
            "File is password-protected (OLE2 EncryptedPackage). Supply the "
            "password or remove protection before uploading."
            if fmt == "encrypted" else
            "Unrecognised file format (not xlsx, xls, html-table or csv).")
    try:
        return _INSPECTORS[fmt](path)
    except UnreadableFile:
        raise
    except Exception as exc:
        raise UnreadableFile(f"Could not inspect file: {type(exc).__name__}: {exc}") from exc
