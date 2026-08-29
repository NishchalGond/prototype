"""Reference data + enrichment.

Source: "Builders data/UAE_Development_Builders.xlsx" — 483 UAE developments
across 7 emirates, 324 with a named master developer, each carrying a
confidence level (High/Medium/Low).

Used to fill `Developer` (only 11% of source files carry one) and to
canonicalise community names. Enrichment never overwrites a value that came
from the source file, and every enriched field is recorded on the record so
the frontend can distinguish sourced data from derived data.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

_PUNCT_RE = re.compile(r"[^a-z0-9]+")
_NOISE = {"the", "at", "residences", "residence", "tower", "towers", "phase",
          "overall", "community", "development"}


def canon(s) -> str:
    if not s:
        return ""
    t = _PUNCT_RE.sub(" ", str(s).lower()).strip()
    words = [w for w in t.split() if w not in _NOISE]
    return " ".join(words) if words else t


@dataclass(frozen=True)
class Development:
    name: str
    emirate: str
    region: str | None
    developer: str | None
    dev_type: str | None
    confidence: str | None


# A canonical key shorter than this is too generic to match on ("hills",
# "park"); requiring length keeps containment from firing on common words.
_MIN_CONTAINMENT_LEN = 5


class ReferenceData:
    def __init__(self, developments: list[Development]):
        self.developments = developments
        self._by_canon: dict[str, Development] = {}
        for d in developments:
            k = canon(d.name)
            if k and (k not in self._by_canon or _rank(d) > _rank(self._by_canon[k])):
                self._by_canon[k] = d
        # Cache is per-instance so a reloaded reference does not serve stale
        # answers, and bounded so a corpus with millions of distinct community
        # spellings cannot grow it without limit.
        self._cache: dict[tuple, Development | None] = {}

    def _match_one(self, k: str) -> Development | None:
        """Longest-prefix, then single-token, match for one canonical key.

        The previous implementation scanned every reference entry for each
        candidate on every row -- 483 comparisons per lookup, several lookups
        per row, tens of millions of rows. It also returned whichever entry
        dict iteration happened to reach first, so "dubai hills estate golf"
        could match a shorter unrelated name before the right one. Walking
        prefixes longest-first is both bounded by the length of the query and
        deterministic: the most specific development wins.
        """
        if not k:
            return None
        if k in self._by_canon:
            return self._by_canon[k]

        words = k.split()
        for i in range(len(words) - 1, 0, -1):
            prefix = " ".join(words[:i])
            if len(prefix) >= _MIN_CONTAINMENT_LEN:
                hit = self._by_canon.get(prefix)
                if hit is not None:
                    return hit

        # No bare single-token fallback. Matching any reference name that
        # happens to appear as a word anywhere in the query assigned "Lime
        # Gardens" (Emaar, Dubai Hills) to Nakheel, because the reference holds
        # a development called "The Gardens". A leading token is already
        # covered by the prefix walk above, which is where a real community
        # name sits; a trailing one is usually a generic word.
        return None

    def lookup(self, *candidates) -> Development | None:
        """Best reference match across the candidate fields, most specific first.

        Candidates are tried in the order given, so callers pass the narrowest
        field (Project) before the broadest (Community).
        """
        keys = tuple(canon(c) for c in candidates if c)
        if not keys:
            return None

        cached = self._cache.get(keys, _MISSING)
        if cached is not _MISSING:
            return cached

        found = None
        # Exact matches on any candidate beat a containment match on an
        # earlier one -- an exact hit on Community is better evidence than a
        # prefix hit on a building name.
        for k in keys:
            if k and k in self._by_canon:
                found = self._by_canon[k]
                break
        else:
            for k in keys:
                found = self._match_one(k)
                if found is not None:
                    break

        if len(self._cache) < 50_000:
            self._cache[keys] = found
        return found

    def __len__(self) -> int:
        return len(self.developments)


_MISSING = object()


_CONF_RANK = {"high": 3, "medium": 2, "low": 1}


def _rank(d: Development) -> int:
    return _CONF_RANK.get((d.confidence or "").lower(), 0) + (2 if d.developer else 0)


def _emirate_of(sheet: str) -> str:
    return sheet.replace("Developers", "").replace("(Verified)", "").strip() or "UAE"


_JSON_FALLBACK = Path(__file__).resolve().parent / "resources" / "uae_developers.json"


def _load_from_json(json_path: Path) -> ReferenceData:
    """Load developments from the bundled JSON (used on Railway where xlsx
    is gitignored)."""
    import json
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    devs = [
        Development(
            name=d["name"],
            emirate=d.get("emirate", "UAE"),
            region=d.get("region"),
            developer=d.get("developer"),
            dev_type=d.get("dev_type"),
            confidence=d.get("confidence"),
        )
        for d in data
    ]
    return ReferenceData(devs)


@lru_cache(maxsize=1)
def load_reference(path_str: str) -> ReferenceData:
    path = Path(path_str)

    # --- if direct json path ---
    if path.suffix.lower() == ".json" and path.exists():
        return _load_from_json(path)

    # --- try xlsx first (local dev) ---
    if path.exists() and path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            pass
        else:
            devs: list[Development] = []
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    if "Developer" not in ws.title:
                        continue
                    rows = ws.iter_rows(values_only=True)
                    try:
                        header = [("" if c is None else str(c)).strip() for c in next(rows)]
                    except StopIteration:
                        continue
                    hl = [h.lower() for h in header]

                    def find(*needles, exact=False):
                        for i, h in enumerate(hl):
                            for n in needles:
                                if (h == n) if exact else (n in h):
                                    return i
                        return None

                    i_name = find("development")
                    i_dev = find("developer", "builder")
                    i_reg = find("region", "zone")
                    i_type = find("development type", "type")
                    i_conf = find("confidence")
                    if i_name is None:
                        continue

                    emirate = _emirate_of(ws.title)
                    for r in rows:
                        if i_name >= len(r) or r[i_name] in (None, ""):
                            continue
                        def get(i):
                            if i is None or i >= len(r):
                                return None
                            v = r[i]
                            if v in (None, "", "—", "-"):
                                return None
                            return str(v).strip()
                        devs.append(Development(
                            name=str(r[i_name]).strip(),
                            emirate=emirate,
                            region=get(i_reg),
                            developer=get(i_dev),
                            dev_type=get(i_type),
                            confidence=get(i_conf),
                        ))
            finally:
                wb.close()
            return ReferenceData(devs)

    # --- fallback: bundled JSON (Railway / production) ---
    if _JSON_FALLBACK.exists():
        return _load_from_json(_JSON_FALLBACK)

    return ReferenceData([])


_MONTHS_AND_NOISE = re.compile(
    r"\b(january|february|march|april|may|june|july|august|september|october|november|december|"
    r"jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|"
    r"201\d|202\d|done|new|full|master|others|contacts|ongoing|update|partial|consolidated|"
    r"file is hanging|hanging|end)\b",
    re.IGNORECASE,
)


def clean_filename_community(filename: str | None) -> str:
    """Extract a clean community / project name from an uploaded file name."""
    if not filename:
        return ""
    stem = Path(filename).stem
    # If filename has a bracketed tag like [Club Villas] Club Villas.xlsx, extract
    bracket_match = re.match(r"^\[(.*?)\]\s*(.*)", stem)
    tag = None
    if bracket_match:
        tag, rest = bracket_match.groups()
        stem = rest if rest.strip() else tag

    # Strip parenthesized tokens like (2022), (c), (1)
    s = re.sub(r"[\(\[\{].*?[\)\]\}]", " ", stem)
    # Strip noise terms
    s = _MONTHS_AND_NOISE.sub(" ", s)
    # Strip punctuation / separators
    s = re.sub(r"[-_./\\]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\b(done|new)\b", "", s, flags=re.IGNORECASE).strip()

    if not s and tag:
        s = _MONTHS_AND_NOISE.sub(" ", tag).strip()

    words = s.split()
    if words and (s.isupper() or s.islower()):
        s = " ".join(w.capitalize() for w in words)
    return s


# Emaar-built sub-developments inside Dubai Hills Estate. Used only as a last
# resort, when neither the source file nor the reference workbook names a
# developer.
#
# Two things this list got wrong before, both worth keeping in mind if it grows:
#
#   * Dubai Hills Estate is a master community, not a single builder's estate.
#     "Ellington House" sits inside it and is built by Ellington Properties, so
#     blanket-stamping Emaar on everything in the estate wrote a factually wrong
#     developer onto those records. Known non-Emaar projects are listed in
#     _DUBAI_HILLS_OTHER_DEVELOPERS below and are matched first.
#   * Matching was a plain substring test, so short entries like "lime" and
#     "collective" fired on any row whose text merely contained those letters.
#     Matching is now on whole words.
_DUBAI_HILLS_SUBPROJECTS = (
    "club villas", "fairway vistas", "sidra", "maple", "park heights",
    "mulberry", "acacia", "park point", "park ridge", "golf place",
    "golf grove", "golf suites", "golf views", "golf ville", "emerald hills",
    "majestic vistas", "hills grove", "hills view", "collective", "elvira",
    "lime gardens", "address hillcrest", "dhe", "dubai hills",
    "parkway vistas", "golf link", "golf palace", "parkridge",
)

# Projects inside Dubai Hills Estate built by someone other than Emaar.
# Checked before the Emaar fallback, so the estate name cannot override a
# builder we actually know.
_DUBAI_HILLS_OTHER_DEVELOPERS = {
    "ellington house": "Ellington Properties",
    "ellington": "Ellington Properties",
}

# The canonical spelling, matching what cleaning.clean_developer emits. Writing
# a second spelling here would split one builder across two facet values.
_DUBAI_HILLS_MASTER_DEVELOPER = "Emaar Properties"


def _mentions(haystack: str, needle: str) -> bool:
    """Whole-word containment, so "lime" does not match inside "Limestone"."""
    return re.search(r"\b%s\b" % re.escape(needle), haystack) is not None


def _alternation(phrases) -> "re.Pattern":
    """One compiled whole-word alternation over many phrases.

    Testing each phrase separately meant ~28 regex searches for every row that
    arrived without a developer. One alternation answers the same question in a
    single pass, and longest-first ordering makes the more specific phrase win
    ("lime gardens" before "lime").
    """
    ordered = sorted(phrases, key=len, reverse=True)
    return re.compile(r"\b(?:%s)\b" % "|".join(re.escape(p) for p in ordered))


_DUBAI_HILLS_EMAAR_RE = _alternation(_DUBAI_HILLS_SUBPROJECTS)
_DUBAI_HILLS_OTHER_RE = _alternation(_DUBAI_HILLS_OTHER_DEVELOPERS)


# Reference entries carry a confidence level. Below this, a developer is not
# written: 9 of the 324 developer-bearing entries are Low, so gating them out
# costs almost no coverage while keeping guesswork out of a field the sales desk
# reads as fact. Medium fills are recorded with a flag so they stay auditable.
MIN_DEVELOPER_CONFIDENCE = "medium"
_CONF_ORDER = {"low": 1, "medium": 2, "high": 3}


def enrich(fields: dict, ref: ReferenceData, source_name: str | None = None,
           flags: list[str] | None = None, properties=None) -> list[str]:
    """Fill Community (from filename if missing) and Developer from the UAE reference workbook.

    Returns the list of field names that were filled by enrichment rather than
    read from the source file. Never overwrites valid source data.

    `flags`, when given, collects provenance notes: which enrichments came from
    a less-than-High-confidence reference entry, and which were withheld. They
    land in the record's validation_flags, so an enriched value can always be
    told apart from a sourced one after the fact.

    `properties`, when given, is a PropertyReference (see
    engine/property_reference.py) consulted for Property Type -- and, when the
    match is to an exact unit rather than a whole building, Bedroom and Size.
    The precision of every such fill is flagged, because a type taken from
    "this tower is all apartments" is a weaker claim than one taken from the
    unit itself, and the desk should be able to tell them apart.
    """
    filled: list[str] = []

    def note(flag: str) -> None:
        if flags is not None and flag not in flags:
            flags.append(flag)

    inferred_comm = clean_filename_community(source_name) if source_name else ""

    # 1. If Community is absent or blank in the row, use the inferred community from filename
    if not fields.get("Community") and inferred_comm:
        fields["Community"] = inferred_comm
        filled.append("community")

    # 2. Check reference table lookup
    if len(ref):
        match = ref.lookup(
            fields.get("Project"),
            fields.get("Sub-Community"),
            fields.get("Community"),
            fields.get("Building/Cluster"),
            inferred_comm if inferred_comm else None,
        )
        if match:
            conf = (match.confidence or "").lower()
            conf_rank = _CONF_ORDER.get(conf, 0)

            if not fields.get("Developer") and match.developer:
                if conf_rank >= _CONF_ORDER[MIN_DEVELOPER_CONFIDENCE]:
                    fields["Developer"] = match.developer
                    filled.append("developer")
                    if conf != "high":
                        note(f"enriched_developer_{conf or 'unrated'}_confidence")
                else:
                    # Recorded rather than silently dropped: a record with no
                    # developer and no explanation looks like missing source
                    # data, when in fact a low-confidence guess was declined.
                    note("enrichment_developer_withheld_low_confidence")

            if not fields.get("Community"):
                # The matched development name is the community. region is the
                # broader area it sits in, so it is only a fallback -- filling
                # Community with a region loses the precision the sales desk
                # filters on.
                community = match.name or match.region
                if community:
                    fields["Community"] = community
                    if "community" not in filled:
                        filled.append("community")
                    note("enriched_community_from_reference")
            # NOTE: match.dev_type is deliberately NOT written to Property
            # Type. Its values are development categories ("Master-planned
            # community", "Rural/desert cadastral zone"), while Property Type
            # means Villa / Apartment / Townhouse. column_mapping.json's
            # do_not_map says the same thing about the source column it came
            # from. Filling it here made records look complete while putting a
            # taxonomy label where the sales desk expects a dwelling type.

    # 3. Property attributes from the property dataset, when the register did
    #    not carry them. Runs after the developer lookup so Community may
    #    already have been filled above and can be matched on.
    if properties is not None and len(properties):
        facts = properties.lookup(fields.get("Community"),
                                  fields.get("Building/Cluster")
                                  or fields.get("Sub-Community"),
                                  fields.get("Unit Number"))
        if facts is not None:
            if not fields.get("Property Type") and facts.property_type:
                fields["Property Type"] = facts.property_type
                filled.append("property_type")
                note(f"enriched_property_type_at_{facts.precision}_precision")
            # Only an exact unit match says anything about a specific
            # property's bedroom count or size. A building- or community-level
            # match is a statement about the group, not about this home.
            if facts.precision == "unit":
                if not fields.get("Bedroom") and facts.bedroom:
                    fields["Bedroom"] = facts.bedroom
                    filled.append("bedroom")
                    note("enriched_bedroom_from_property_reference")
                if not fields.get("Size") and facts.size:
                    fields["Size"] = facts.size
                    filled.append("size")
                    note("enriched_size_from_property_reference")

    # 4. Dubai Hills Master Builder Fallback (All Dubai Hills sub-communities are Emaar)
    all_text = " ".join(
        str(fields.get(k) or "") for k in ("Community", "Sub-Community", "Project", "Building/Cluster")
    ).lower() + " " + inferred_comm.lower()

    if not fields.get("Developer"):
        # A named non-Emaar builder in the estate wins over the master-developer
        # fallback: Dubai Hills Estate is a master community, not one builder's
        # estate, and stamping Emaar on a plot someone else built is a factual
        # error the sales desk has no way to spot.
        other = _DUBAI_HILLS_OTHER_RE.search(all_text)
        if other:
            fields["Developer"] = _DUBAI_HILLS_OTHER_DEVELOPERS[other.group(0)]
            filled.append("developer")
            note("enriched_developer_from_project_map")
        elif _DUBAI_HILLS_EMAAR_RE.search(all_text):
            fields["Developer"] = _DUBAI_HILLS_MASTER_DEVELOPER
            filled.append("developer")
            note("enriched_developer_from_master_community")

    return filled
