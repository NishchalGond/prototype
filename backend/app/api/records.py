"""Record search / filter / detail + dashboard stats + mapping introspection."""
from __future__ import annotations

import json
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from sqlalchemy import and_, func, or_, select, text as sa_text
from sqlalchemy.orm import Session, selectinload

from ..config import settings
from ..database.session import IS_POSTGRES, get_db, get_read_db
from ..core.search import build_search_filter
from ..core.security import (
    get_current_user, require_export_permission, require_role,
)
from ..models.models import (
    ExportAuditLog, Lead, LeadStage, ProcessingError, ProcessingJob, Record,
    RecordEditAudit, RecordStatus, SourceFile, User, UserRole,
)
from ..schemas.schemas import (
    AliasRequest, ColumnMappingOut, DashboardStats, FilterOptions, JobOut, Page, RecordOut,
)
from engine import cleaning as C
from engine import validation as V

router = APIRouter()

SORTABLE = {
    "id": Record.id, "name": Record.name, "community": Record.community,
    "sub_community": Record.sub_community, "building_cluster": Record.building_cluster,
    "unit_number": Record.unit_number, "size": Record.size,
    "bedroom": Record.bedroom, "procedure_value": Record.procedure_value,
    "mobile_1": Record.mobile_1, "status": Record.status,
    "created_at": Record.created_at, "record_date": Record.record_date,
}

# Free-text search now lives in core/search.py. It tokenises the query and
# requires every token to match, so "Mohammed Ahmed Marina Heights" finds the
# owner even though no single column holds all four words -- the old whole-
# phrase ILIKE across 13 columns could not, and forced a sequential scan
# besides. See that module for why the column list moved with it.

# Counting every matching row costs a full scan of the match set, which at 20M
# rows is seconds for a broad filter and is spent on a number nobody reads past
# the first significant digit. Counting stops here and the UI shows "20,000+".
COUNT_CEILING = 20_000


def _build_records_query(
    q: str | None = None,
    community: str | None = None,
    sub_community: str | None = None,
    building_cluster: str | None = None,
    property_type: str | None = None,
    bedroom: str | None = None,
    developer: str | None = None,
    nationality: str | None = None,
    source_file: str | None = None,
    job_id: int | None = None,
    record_status: str | None = None,
    has_mobile: bool | None = None,
    has_email: bool | None = None,
):
    stmt = select(Record)
    search = build_search_filter(q, is_postgres=IS_POSTGRES)
    if search is not None:
        stmt = stmt.where(search)

    for col, val in (
        (Record.community, community), (Record.sub_community, sub_community),
        (Record.building_cluster, building_cluster),
        (Record.developer, developer),
        (Record.nationality, nationality), (Record.source_file, source_file),
    ):
        if val:
            stmt = stmt.where(col == val)

    if bedroom:
        b_clean = bedroom.strip()
        stmt = stmt.where(
            or_(
                Record.bedroom == b_clean,
                Record.bedroom.ilike(f"%{b_clean}%")
            )
        )

    # "Verified valid mobile": non-null, non-N/A, and matching one of the three
    # accepted E.164 shapes (UAE mobile, UAE landline, other international) --
    # so truncated junk like +55240883 or 055240883 is excluded.
    #
    # This used to be three regexes evaluated against every row on every page
    # load, which no index can serve and which the default view runs
    # unconditionally. The rule now lives in the has_valid_mobile generated
    # column, computed once at write time, and the partial indexes added in
    # 9c41ab7de205 are built on it.
    valid_mobile_filter = Record.has_valid_mobile.is_(True)

    if record_status:
        st_upper = record_status.upper()
        if st_upper in ("ALL", "ALL_RECORDS", "ALL_WITH_INCOMPLETE", "SHOW_ALL"):
            pass  # show all records including duplicates and incomplete
        elif st_upper in ("INCOMPLETE", "MISSING_CONTACT"):
            stmt = stmt.where(
                or_(
                    Record.status == "INCOMPLETE",
                    Record.mobile_1.is_(None),
                    ~valid_mobile_filter,
                )
            )
        elif st_upper in ("VALID", "COMPLETE"):
            # All valid records (outreach-ready with standard valid phone)
            stmt = stmt.where(Record.status == "VALID")
            stmt = stmt.where(valid_mobile_filter)
        else:
            stmt = stmt.where(Record.status == st_upper)
    else:
        # Default: show all valid outreach-ready records (with verified valid phone)
        stmt = stmt.where(Record.status == "VALID")
        stmt = stmt.where(valid_mobile_filter)

    if property_type:
        stmt = stmt.where(Record.property_type.ilike(property_type))

    if job_id is not None:
        stmt = stmt.where(Record.job_id == job_id)
    if has_mobile is True:
        stmt = stmt.where(Record.mobile_1.is_not(None))
    elif has_mobile is False:
        stmt = stmt.where(Record.mobile_1.is_(None))
    if has_email is True:
        stmt = stmt.where(Record.email_address.is_not(None))
    elif has_email is False:
        stmt = stmt.where(Record.email_address.is_(None))

    # Opt-outs, enforced here rather than at each call site, because the two
    # call sites are exactly the paths that must honour them: the list the desk
    # calls from, and the export it takes off-platform. A DO_NOT_CONTACT stage
    # that still appears in either is not an opt-out, it is a note.
    #
    # An anti-join on identity_hash rather than record_id: the lead survives
    # reprocessing and its record_id is briefly NULL afterwards, so keying on
    # the pointer would let an opted-out person reappear in the window between
    # a reprocess and the relink.
    #
    # leads is small -- a row exists only where someone was actually worked --
    # so this stays a cheap semi-join against a unique index, not a scan.
    stmt = stmt.where(
        ~select(Lead.id)
        .where(Lead.identity_hash == Record.identity_hash,
               Lead.stage == LeadStage.DO_NOT_CONTACT)
        .exists()
    )

    return stmt


@router.get("/records", response_model=Page[RecordOut])
def list_records(
    q: str | None = Query(None, description="Free-text search across name, location, contact."),
    community: str | None = None,
    sub_community: str | None = None,
    building_cluster: str | None = None,
    property_type: str | None = None,
    bedroom: str | None = None,
    developer: str | None = None,
    nationality: str | None = None,
    source_file: str | None = None,
    job_id: int | None = None,
    status: str | None = None,
    record_status: str | None = None,
    has_mobile: bool | None = None,
    has_email: bool | None = None,
    sort_by: str = Query("id"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE),
    _user: User = Depends(get_current_user),
    db: Session = Depends(get_read_db),
):
    if sort_by not in SORTABLE:
        raise HTTPException(400, f"sort_by must be one of {sorted(SORTABLE)}")

    effective_status = status or record_status
    stmt = _build_records_query(
        q=q, community=community, sub_community=sub_community,
        building_cluster=building_cluster, property_type=property_type,
        bedroom=bedroom, developer=developer, nationality=nationality,
        source_file=source_file, job_id=job_id, record_status=effective_status,
        has_mobile=has_mobile, has_email=has_email,
    )

    # Counting the full match set is a scan of every matching row. On a broad
    # filter over 20M records that is the slowest part of the request, and it
    # produces a number the UI only uses to draw a page count. Counting stops at
    # COUNT_CEILING: below it the total is exact, at it the UI shows "20,000+"
    # and the user narrows their filter, which is what they should do anyway.
    total = db.scalar(
        select(func.count()).select_from(stmt.limit(COUNT_CEILING).subquery())
    ) or 0
    total_capped = total >= COUNT_CEILING

    col = SORTABLE[sort_by]
    if sort_by == "name" and sort_dir == "asc":
        # On default initial page load, prioritize complete records where procedure_value > 0 so VALUE (AED) and BEDROOM are visible right at the top
        stmt = stmt.order_by(
            Record.procedure_value.desc().nullslast(),
            Record.bedroom.desc().nullslast(),
            col.asc().nullslast(),
            Record.id.desc(),
        )
    else:
        order_clause = col.desc().nullslast() if sort_dir == "desc" else col.asc().nullslast()
        stmt = stmt.order_by(order_clause, Record.id.desc())
    # OFFSET is bounded by COUNT_CEILING above (the UI cannot page past the
    # capped total), so the planner never walks more than a few thousand index
    # entries before the LIMIT. That keeps plain offset pagination viable; if
    # deeper navigation is ever exposed, this is the point to switch to a
    # keyset cursor on (sort_key, id).
    rows = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    pages = (total + page_size - 1) // page_size
    return Page[RecordOut](
        items=[RecordOut.model_validate(r) for r in rows], total=total, page=page,
        page_size=page_size, total_pages=pages, has_next=page < pages, has_prev=page > 1,
        total_capped=total_capped,
    )


@router.get("/records/export")
def export_records(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    q: str | None = None,
    community: str | None = None,
    sub_community: str | None = None,
    building_cluster: str | None = None,
    property_type: str | None = None,
    bedroom: str | None = None,
    developer: str | None = None,
    nationality: str | None = None,
    source_file: str | None = None,
    job_id: int | None = None,
    status: str | None = None,
    record_status: str | None = None,
    has_mobile: bool | None = None,
    has_email: bool | None = None,
    sort_by: str = Query("id"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    limit: int = Query(50000, ge=1, le=100000),
    request: Request = None,
    current_user: User = Depends(require_export_permission),
    db: Session = Depends(get_db),
):
    """Export filtered dataset to CSV or Excel (.xlsx). Exactly respects active search and filters."""
    import csv
    import io
    from datetime import datetime, timezone
    from fastapi.responses import StreamingResponse

    if sort_by not in SORTABLE:
        sort_by = "id"

    effective_status = status or record_status
    stmt = _build_records_query(
        q=q, community=community, sub_community=sub_community,
        building_cluster=building_cluster, property_type=property_type,
        bedroom=bedroom, developer=developer, nationality=nationality,
        source_file=source_file, job_id=job_id, record_status=effective_status,
        has_mobile=has_mobile, has_email=has_email,
    )

    col = SORTABLE[sort_by]
    order_clause = col.desc().nullslast() if sort_dir == "desc" else col.asc().nullslast()
    stmt = stmt.order_by(order_clause, Record.id.desc()).limit(limit)

    # yield_per streams the result in chunks instead of materialising up to
    # `limit` (100k) ORM objects at once, which was hundreds of MB of resident
    # memory per concurrent export and the most likely cause of an OOM kill.
    def iter_rows():
        yield from db.scalars(stmt.execution_options(yield_per=1000))

    headers = [
        "Record ID", "Name", "Community", "Sub-Community", "Building / Cluster",
        "Unit Number", "Plot Number", "Plot Reg. No", "DMNO", "DMsubno",
        "Bedroom", "Property Type", "Developer", "Project", "Party Type (Buyer/Seller)",
        "Size (Sq.Ft)", "Procedure Value (AED)",
        "Mobile 1 (Primary)", "Mobile 2", "Mobile 3", "Email Address",
        "Nationality", "PI Number", "Status", "Source File", "Record Date",
    ]

    def extract_row_values(r: Record) -> list:
        return [
            r.id,
            r.name or "",
            r.community or "",
            r.sub_community or "",
            r.building_cluster or "",
            r.unit_number or "",
            r.plot_number or "",
            r.plot_reg_no or "",
            r.dmno or "",
            r.dmsubno or "",
            r.bedroom or "",
            r.property_type or "",
            r.developer or "",
            r.project or "",
            r.party_type or "",
            r.size if r.size is not None else "",
            r.procedure_value if r.procedure_value is not None else "",
            r.mobile_1 or "",
            r.mobile_2 or "",
            r.mobile_3 or "",
            r.email_address or "",
            r.nationality or "",
            r.pi_number or "",
            r.status or "",
            r.source_file or "",
            r.record_date.strftime("%Y-%m-%d") if r.record_date else "",
        ]

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    # Record export audit log
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    user_id = current_user.id
    user_email = current_user.email

    # Written before a single byte leaves, so an aborted or interrupted download
    # still leaves a record that this data was requested. The row count comes
    # from a COUNT over the same filtered statement, since the rows themselves
    # are now streamed rather than materialised.
    row_count = db.scalar(
        select(func.count()).select_from(stmt.order_by(None).subquery())) or 0

    audit_entry = ExportAuditLog(
        user_id=user_id,
        user_email=user_email,
        format=format.upper(),
        filter_criteria={
            "q": q, "community": community, "property_type": property_type,
            "bedroom": bedroom, "developer": developer, "status": effective_status,
        },
        row_count=row_count,
        ip_address=client_ip,
        user_agent=user_agent,
    )
    db.add(audit_entry)
    db.commit()

    if format == "xlsx":
        import openpyxl
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # write_only holds one row at a time and spools the rest to a temp file.
        # An xlsx is a zip archive so it must still be complete before sending,
        # but peak memory no longer scales with the export size.
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Datalink Export")

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        # In write_only mode column widths must be set before any row is
        # appended, so they are derived from the header text instead of from a
        # post-hoc scan of every cell. Same 10..45 clamp as before.
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 3, 10), 45)
        ws.freeze_panes = "A2"

        styled_header = []
        for h in headers:
            cell = WriteOnlyCell(ws, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            styled_header.append(cell)
        ws.append(styled_header)

        for r in iter_rows():
            ws.append(extract_row_values(r))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"datalink_records_{stamp}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    else:
        # Streamed a chunk at a time: memory stays flat regardless of row count
        # and the download starts immediately instead of after a full build.
        # Byte-for-byte the same output as before (BOM, header, same rows).
        def csv_chunks():
            buf = io.StringIO()
            buf.write("\ufeff")  # UTF-8 BOM, keeps Excel happy with UTF-8
            writer = csv.writer(buf)
            writer.writerow(headers)
            for r in iter_rows():
                writer.writerow(extract_row_values(r))
                if buf.tell() > 64 * 1024:
                    yield buf.getvalue().encode("utf-8")
                    buf.seek(0)
                    buf.truncate(0)
            if buf.tell():
                yield buf.getvalue().encode("utf-8")

        filename = f"datalink_records_{stamp}.csv"
        return StreamingResponse(
            csv_chunks(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


# --------------------------------------------------------------------------
# Cached aggregate access.
#
# Both helpers read a materialised view and return None if it is not there, so
# every caller keeps a live-query fallback. That matters in three situations:
# SQLite dev databases (no materialised views at all), a deploy where the app
# rolls out before `alembic upgrade head` has finished, and a view that has been
# dropped by hand. In all three the dashboard stays correct and merely slow,
# rather than erroring.
def _matview(db: Session, sql: str):
    try:
        return db.execute(sa_text(sql)).all()
    except Exception:
        # A missing relation aborts the surrounding PostgreSQL transaction, so
        # the session must be rolled back before the fallback query can run on it.
        db.rollback()
        return None


def _facet_cache(db: Session) -> dict[str, list[str]] | None:
    """Return {column_name: [distinct values]} from mv_record_facets, or None."""
    rows = _matview(
        db,
        "SELECT field, value FROM mv_record_facets "
        "WHERE value <> '' ORDER BY field, value",
    )
    if rows is None:
        return None
    out: dict[str, list[str]] = {}
    for field, value in rows:
        out.setdefault(field, []).append(value)
    return out


@router.get("/records/filters", response_model=FilterOptions)
def filter_options(
    _user: User = Depends(get_current_user),
    db: Session = Depends(get_read_db),
):
    """Distinct values for the dashboard filter dropdowns.

    Served from the mv_record_facets materialised view rather than seven
    SELECT DISTINCT scans of the records table. Dropdown contents only change
    when a file is ingested, so they are refreshed on that event (and on a
    schedule) instead of being recomputed for every dashboard load by every one
    of ~60 users. Falls back to live DISTINCT when the view is absent, which is
    the case on SQLite dev databases and before the migration has run.
    """
    facets = _facet_cache(db)

    def distinct(col, limit=500, is_community=False):
        cached = facets.get(col.key) if facets is not None else None
        if cached is not None:
            raw_vals = cached[:limit]
        else:
            raw_vals = [v for (v,) in db.execute(
                select(col).where(col.is_not(None)).distinct().order_by(col).limit(limit)
            ).all() if v]
        if is_community:
            valid_comms = []
            for v in raw_vals:
                s_low = str(v).lower()
                if "total owner" in s_low or "owner detail" in s_low or "owners data" in s_low:
                    continue
                cleaned = C.clean_community(v)
                if cleaned and "total owner" not in cleaned.lower() and cleaned not in valid_comms:
                    valid_comms.append(cleaned)
            return sorted(valid_comms)
        return raw_vals

    return FilterOptions(
        communities=distinct(Record.community, is_community=True),
        sub_communities=distinct(Record.sub_community),
        property_types=distinct(Record.property_type),
        bedrooms=distinct(Record.bedroom),
        developers=distinct(Record.developer),
        source_files=distinct(Record.source_file),
        statuses=distinct(Record.status),
    )


from ..schemas.schemas import (
    ColumnMappingOut, DashboardStats, FilterOptions, JobOut, Page, RecordOut, RecordUpdate,
)

@router.get("/records/{record_id}", response_model=RecordOut)
def get_record(
    record_id: int,
    _user: User = Depends(get_current_user),
    db: Session = Depends(get_read_db),
):
    rec = db.get(Record, record_id)
    if not rec:
        raise HTTPException(404, f"Record {record_id} not found.")
    return RecordOut.model_validate(rec)


@router.get("/records/{record_id}/audits")
def get_record_audits(
    record_id: int,
    _user: User = Depends(get_current_user),
    db: Session = Depends(get_read_db),
):
    """Retrieve complete audit history for manual edits to this record."""
    audits = db.scalars(
        select(RecordEditAudit)
        .where(RecordEditAudit.record_id == record_id)
        .order_by(RecordEditAudit.edited_at.desc())
    ).all()
    return [
        {
            "id": a.id,
            "field_name": a.field_name,
            "old_value": a.old_value,
            "new_value": a.new_value,
            "user_email": a.user_email,
            "edited_at": a.edited_at.isoformat(),
        }
        for a in audits
    ]


@router.put("/records/{record_id}", response_model=RecordOut)
def update_record(
    record_id: int,
    body: RecordUpdate,
    request: Request,
    current_user: User = Depends(
        require_role([UserRole.ADMIN, UserRole.DATA_PROCESSOR])),
    db: Session = Depends(get_db)
):
    """Hardened update endpoint: cleans fields, re-validates, re-hashes, and logs audit trail."""
    rec = db.get(Record, record_id)
    if not rec:
        raise HTTPException(404, f"Record {record_id} not found.")

    user_id = current_user.id
    user_email = current_user.email

    update_data = body.model_dump(exclude_unset=True)
    
    # Whitelist of editable business fields
    EDITABLE_FIELDS = {
        "name", "community", "sub_community", "building_cluster", "unit_number",
        "plot_number", "plot_reg_no", "dmno", "dmsubno", "bedroom", "party_type",
        "mobile_1", "mobile_2", "mobile_3", "email_address", "pi_number",
        "nationality", "property_type", "procedure_value", "size", "developer", "project"
    }

    changes: list[RecordEditAudit] = []

    for field, value in update_data.items():
        if field in EDITABLE_FIELDS:
            # Clean individual field with engine cleaning rules
            cleaned_val = value
            if field == "name":
                cleaned_val = C.clean_name(value)
            elif field == "community":
                cleaned_val = C.clean_community(value)
            elif field in ("sub_community", "building_cluster", "dmno", "dmsubno", "pi_number", "property_type", "developer", "project", "plot_reg_no"):
                cleaned_val = C.clean_text(value)
            elif field in ("unit_number", "plot_number"):
                cleaned_val = C.clean_unit(value)
            elif field in ("mobile_1", "mobile_2", "mobile_3"):
                cleaned_val, _ = C.clean_phone(value)
            elif field == "email_address":
                cleaned_val, _ = C.clean_email(value)
            elif field == "bedroom":
                cleaned_val, _ = C.clean_bedroom(value)
            elif field == "party_type":
                cleaned_val = C.clean_party_type(value)
            elif field == "nationality":
                cleaned_val = C.clean_nationality(value)
            elif field == "procedure_value":
                cleaned_val = C.clean_number(value)
            elif field == "size":
                cleaned_val = C.clean_size(value)

            old_val = getattr(rec, field)
            if old_val != cleaned_val:
                changes.append(
                    RecordEditAudit(
                        record_id=rec.id,
                        user_id=user_id,
                        user_email=user_email,
                        field_name=field,
                        old_value=str(old_val) if old_val is not None else None,
                        new_value=str(cleaned_val) if cleaned_val is not None else None,
                    )
                )
                setattr(rec, field, cleaned_val)

    # Re-validate status
    row_dict = {
        "name": rec.name,
        "mobile_1": rec.mobile_1,
        "mobile_2": rec.mobile_2,
        "mobile_3": rec.mobile_3,
        "email_address": rec.email_address,
        "community": rec.community,
        "sub_community": rec.sub_community,
        "building_cluster": rec.building_cluster,
        "unit_number": rec.unit_number,
        "plot_number": rec.plot_number,
        "pi_number": rec.pi_number,
        "developer": rec.developer,
        "project": rec.project,
        "bedroom": rec.bedroom,
        "procedure_value": rec.procedure_value,
        "property_type": rec.property_type,
        "party_type": rec.party_type,
    }
    is_valid, flags = V.validate(row_dict)

    has_property = V.is_valid_property_context(row_dict)
    has_contact = bool(rec.mobile_1 or rec.email_address)
    has_name = bool(rec.name and str(rec.name).strip())

    if not is_valid:
        rec.status = RecordStatus.INVALID
    elif has_name and has_contact and has_property:
        rec.status = RecordStatus.VALID
    else:
        rec.status = "INCOMPLETE"

    rec.identity_hash = V.identity_hash(row_dict)
    rec.validation_flags = flags

    for audit in changes:
        db.add(audit)

    db.commit()
    db.refresh(rec)
    return RecordOut.model_validate(rec)


# --------------------------------------------------------------------------
@router.get("/dashboard/stats", response_model=DashboardStats)
def dashboard_stats(
    _user: User = Depends(get_current_user),
    db: Session = Depends(get_read_db),
):
    # The stat tiles are polled by every open dashboard. Computing them live
    # meant a full-table COUNT plus a GROUP BY plus a 17-aggregate scan per
    # poll, per user -- at 60 users that is continuous full scans of a 20M-row
    # table for numbers that change only when a job finishes. mv_record_stats
    # holds all of it in a single row; the live path below still runs when the
    # view is unavailable.
    cached = _matview(db, "SELECT * FROM mv_record_stats LIMIT 1")
    stats_row = cached[0]._mapping if cached else None

    if stats_row is not None:
        total_records = stats_row["total_records"] or 0
        by_status = {
            RecordStatus.VALID: stats_row["valid_records"] or 0,
            RecordStatus.INVALID: stats_row["invalid_records"] or 0,
            "DUPLICATE": stats_row["duplicate_records"] or 0,
        }
    else:
        total_records = db.scalar(select(func.count(Record.id))) or 0
        by_status = dict(db.execute(
            select(Record.status, func.count(Record.id)).group_by(Record.status)).all())

    jobs_by_status = dict(db.execute(
        select(ProcessingJob.status, func.count(ProcessingJob.id))
        .group_by(ProcessingJob.status)).all())

    # mv_record_facets already stores a per-community count, so the top-10 chart
    # is a 10-row read from a small view instead of a GROUP BY over every record.
    top_rows = _matview(
        db,
        "SELECT value, n FROM mv_record_facets WHERE field = 'community' "
        "ORDER BY n DESC LIMIT 10",
    )
    if top_rows is None:
        top_rows = db.execute(
            select(Record.community, func.count(Record.id))
            .where(Record.community.is_not(None))
            .group_by(Record.community).order_by(func.count(Record.id).desc()).limit(10)
        ).all()
    top = [{"community": c, "count": n} for c, n in top_rows]

    # Field completeness in a single pass. This used to issue one
    # `COUNT(*) WHERE col IS NOT NULL` per field -- 17 sequential scans of the
    # whole records table on every dashboard poll. count(col) already ignores
    # NULLs, so all 17 collapse into one scan with 17 aggregates.
    COMPLETENESS_FIELDS = (
        ("name", Record.name), ("community", Record.community),
        ("sub_community", Record.sub_community),
        ("building_cluster", Record.building_cluster),
        ("unit_number", Record.unit_number), ("size", Record.size),
        ("bedroom", Record.bedroom), ("mobile_1", Record.mobile_1),
        ("email_address", Record.email_address), ("developer", Record.developer),
        ("project", Record.project), ("nationality", Record.nationality),
        ("property_type", Record.property_type), ("record_date", Record.record_date),
        ("procedure_value", Record.procedure_value),
        ("party_type", Record.party_type), ("pi_number", Record.pi_number),
    )

    completeness: dict[str, float] = {}
    if total_records:
        if stats_row is not None:
            # Same 17 aggregates, already computed by the materialised view.
            counts = [stats_row[f"c_{label}"] or 0
                      for label, _ in COMPLETENESS_FIELDS]
        else:
            counts = db.execute(
                select(*[func.count(col) for _, col in COMPLETENESS_FIELDS])
            ).one()
        completeness = {
            label: round(100.0 * n / total_records, 1)
            for (label, _), n in zip(COMPLETENESS_FIELDS, counts)
        }

    # selectinload: _job_out/JobOut reads job.source_file.filename, which
    # lazy-loads one SELECT per job without this.
    recent = db.scalars(
        select(ProcessingJob)
        .options(selectinload(ProcessingJob.source_file))
        .order_by(ProcessingJob.id.desc()).limit(10)).all()
    recent_out: list[JobOut] = []
    for j in recent:
        o = JobOut.model_validate(j)
        o.filename = j.source_file.filename if j.source_file else None
        recent_out.append(o)
    last_out = recent_out[0] if recent_out else None

    valid = by_status.get(RecordStatus.VALID, 0)
    success_rate = round(100.0 * valid / total_records, 1) if total_records else 0.0

    return DashboardStats(
        success_rate=success_rate,
        community_distribution=[{"name": c["community"], "count": c["count"]}
                                for c in top],
        recent_jobs=recent_out,
        total_files=db.scalar(select(func.count(SourceFile.id))) or 0,
        total_jobs=db.scalar(select(func.count(ProcessingJob.id))) or 0,
        total_records=total_records,
        valid_records=by_status.get(RecordStatus.VALID, 0),
        invalid_records=by_status.get(RecordStatus.INVALID, 0),
        duplicate_records=db.scalar(
            select(func.coalesce(func.sum(ProcessingJob.duplicate_rows), 0))) or 0,
        total_errors=db.scalar(select(func.count(ProcessingError.id))) or 0,
        jobs_by_status=jobs_by_status,
        records_by_status=by_status,
        top_communities=top,
        field_completeness=completeness,
        last_job=last_out,
    )


@router.get("/column-mappings", response_model=ColumnMappingOut)
def column_mappings(_user: User = Depends(get_current_user)):
    """Expose the mapping layer so the dashboard can show why a column landed where."""
    from engine.mapping import FIELD_TO_COLUMN

    path = Path(__file__).resolve().parents[3] / "engine" / "resources" / "column_mapping.json"
    cfg = json.loads(path.read_text(encoding="utf8"))
    return ColumnMappingOut(
        target_fields=cfg["target_fields"],
        field_to_column=FIELD_TO_COLUMN,
        aliases=cfg["aliases"],
        composite_fields=cfg.get("composite_fields", {}),
        exclude_columns=cfg.get("exclude_columns", []),
        do_not_map=cfg.get("do_not_map", {}),
        alias_count=sum(len(v) for v in cfg["aliases"].values()),
    )


def _sync_alias_files(cfg: dict) -> None:
    """Save updated mapping config to disk and reload in-memory engine structures."""
    root_path = Path(__file__).resolve().parents[3] / "column_mapping.json"
    engine_path = Path(__file__).resolve().parents[3] / "engine" / "resources" / "column_mapping.json"

    formatted = json.dumps(cfg, indent=2, ensure_ascii=False)
    root_path.write_text(formatted, encoding="utf8")
    engine_path.write_text(formatted, encoding="utf8")

    # Reload engine.mapping in memory dynamically
    import engine.mapping
    engine.mapping._CFG = cfg
    engine.mapping.ALIAS = {}
    for _t, _srcs in cfg["aliases"].items():
        for _src in _srcs:
            _k = engine.mapping.norm_header(_src)
            engine.mapping.ALIAS.setdefault(_k, []).append(_t)


@router.post("/column-mappings/alias", response_model=ColumnMappingOut)
def add_alias(
    body: AliasRequest,
    _user: User = Depends(require_role([UserRole.ADMIN])),
):
    """Add a new custom header alias for a target field and persist permanently."""
    path = Path(__file__).resolve().parents[3] / "engine" / "resources" / "column_mapping.json"
    cfg = json.loads(path.read_text(encoding="utf8"))

    target = body.target_field.strip()
    alias_str = body.alias.strip()
    if not target or not alias_str:
        raise HTTPException(400, "Both target_field and alias must be non-empty strings.")

    if target not in cfg["target_fields"] and target not in cfg["aliases"]:
        raise HTTPException(404, f"Target field '{target}' not found in canonical target list.")

    aliases_list = cfg["aliases"].setdefault(target, [])
    # Case-insensitive duplicate check
    if not any(a.lower() == alias_str.lower() for a in aliases_list):
        aliases_list.append(alias_str)

    _sync_alias_files(cfg)
    return column_mappings()


@router.delete("/column-mappings/alias", response_model=ColumnMappingOut)
def remove_alias(
    body: AliasRequest,
    _user: User = Depends(require_role([UserRole.ADMIN])),
):
    """Remove a custom header alias permanently."""
    path = Path(__file__).resolve().parents[3] / "engine" / "resources" / "column_mapping.json"
    cfg = json.loads(path.read_text(encoding="utf8"))

    target = body.target_field.strip()
    alias_str = body.alias.strip()

    if target in cfg["aliases"]:
        cfg["aliases"][target] = [a for a in cfg["aliases"][target] if a.lower() != alias_str.lower()]

    _sync_alias_files(cfg)
    return column_mappings()

